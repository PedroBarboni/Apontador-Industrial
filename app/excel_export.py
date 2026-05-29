"""
excel_export.py - Exportação no padrão completo da aba BD_LANÇAMENTOS

Gera a planilha no mesmo formato da base usada no Power BI, com TODAS as colunas
A até T do modelo enviado:

A  MAQUINA
B  PRODUTO
C  OPERADOR
D  QTD PRODUZIDA
E  HORA INÍCIO
F  HORA FIM
G  HORAS PARADAS
H  HORAS TRABALHADAS
I  CAP. PRODUÇÃO (Hrs trabalhadas)
J  PRODUT. %
K  TOTAL HORA
L  HORA/DIA
M  MÊS
N  DATA
O  OPERAÇÃO
P  MOTIVO PARADA
Q  HORAS "ÚTEIS"
R  HORAS NO DIA
S  HORA/DIA2
T  CAP. PRODUÇÃO (Turno completo)

Regra do dia produtivo:
- Turno B: exibido visualmente como 21:00 até 06:00, usando somente registros da data selecionada
- Turno A: 06:00 até 16:00 da data selecionada
- A planilha sempre imprime primeiro o Turno B e depois o Turno A.
"""

from __future__ import annotations

from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise ImportError("Biblioteca openpyxl não encontrada. Instale com: pip install openpyxl") from exc

from app import database as db
from app.utils import tempo_para_minutos


META_DIA_KG = 1700

# Cores do modelo
AZUL_CLARO = "BDD7EE"
CINZA_ESCURO = "595959"
CINZA = "D9D9D9"
BRANCO = "FFFFFF"
PRETO = "000000"
VERDE_CLARO = "C6EFCE"
VERDE_TEXTO = "008000"
VERMELHO_CLARO = "FFC7CE"
VERMELHO_TEXTO = "9C0006"
AMARELO = "FFFF00"
LARANJA = "F4B183"
MARROM = "C65911"

BORDA_PRETA = Border(
    left=Side(style="thin", color=PRETO),
    right=Side(style="thin", color=PRETO),
    top=Side(style="thin", color=PRETO),
    bottom=Side(style="thin", color=PRETO),
)

BORDA_MEDIA = Border(
    left=Side(style="medium", color=PRETO),
    right=Side(style="medium", color=PRETO),
    top=Side(style="medium", color=PRETO),
    bottom=Side(style="medium", color=PRETO),
)


def _fill(cor: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor)


def _font(bold: bool = False, color: str = PRETO, size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _normalizar(texto: str) -> str:
    return str(texto or "").upper().strip().replace(".", "")


def _parse_data(data_iso: str) -> date:
    return datetime.strptime(str(data_iso), "%Y-%m-%d").date()


def _parse_hora(hora_txt: str) -> time:
    try:
        return datetime.strptime(str(hora_txt or "00:00").strip(), "%H:%M").time()
    except Exception:
        return time(0, 0)


def _datetime_apontamento(data_iso: str, hora_txt: str) -> datetime:
    return datetime.combine(_parse_data(data_iso), _parse_hora(hora_txt))


def _datetime_inicio_apontamento(apontamento: dict) -> datetime:
    data_base = apontamento.get("data") or apontamento.get("data_iso") or datetime.now().strftime("%Y-%m-%d")
    return _datetime_apontamento(str(data_base), str(apontamento.get("hora_inicio") or "00:00"))


def _datetime_fim_apontamento(apontamento: dict) -> datetime:
    data_base = apontamento.get("data") or apontamento.get("data_iso") or datetime.now().strftime("%Y-%m-%d")
    inicio = _datetime_apontamento(str(data_base), str(apontamento.get("hora_inicio") or "00:00"))
    fim = _datetime_apontamento(str(data_base), str(apontamento.get("hora_fim") or "00:00"))
    if fim < inicio:
        fim += timedelta(days=1)
    return fim


def _tempo_excel(tempo_txt: str) -> float:
    return tempo_para_minutos(tempo_txt or "00:00") / 1440


def _minutos_para_excel(minutos: float) -> float:
    return float(minutos or 0) / 1440


def _peso_unitario(produto: str) -> float:
    try:
        info = db.buscar_produto_por_nome(str(produto or "").strip().upper())
        if info:
            return float(info.get("peso_unitario") or 0)
    except Exception:
        pass
    return 0.0


def _prod_hora(produto: str) -> float:
    try:
        info = db.buscar_produto_por_nome(str(produto or "").strip().upper())
        if info:
            return float(info.get("prod_hora") or 0)
    except Exception:
        pass
    return 0.0


def _safe_nome_arquivo(texto: str) -> str:
    for ch in '<>:"/\\|?*':
        texto = texto.replace(ch, "-")
    return texto


def _eh_pre_acabado(apontamento: dict) -> bool:
    """Regra de pré-acabados solicitada."""
    maquina = _normalizar(apontamento.get("maquina") or "")
    produto = _normalizar(apontamento.get("produto") or "")
    processo = _normalizar(apontamento.get("processo") or apontamento.get("operacao") or "")

    # Peças produzidas na Solda 1 e Solda 2
    if maquina in ("SOLDA 1", "SOLDA 2", "SOLDA1", "SOLDA2"):
        return True

    # Laminadora laminando P. Cabo ou Pino Condutor
    eh_laminadora = maquina.startswith("LAMI") or "LAMINADORA" in maquina
    eh_p_cabo = "P CABO" in produto or "PRENSA CABO" in produto or "P CABO" in processo or "PRENSA CABO" in processo
    eh_pino = "PINO CONDUTOR" in produto or "PINO CONDUTOR" in processo
    eh_laminacao = "LAMIN" in produto or "LAMIN" in processo
    if eh_laminadora and eh_laminacao and (eh_p_cabo or eh_pino):
        return True

    # CNC usinando Terminal AT
    eh_cnc = maquina.startswith("CNC")
    eh_terminal_at = "TERMINAL AT" in produto or "TERMINAL AT" in processo
    eh_usinagem = "USINAGEM" in produto or "USINAGEM" in processo
    if eh_cnc and eh_terminal_at and eh_usinagem:
        return True

    return False


def _identificar_turno_export(apontamento: dict) -> Optional[str]:
    """Identifica o turno do registro para a exportação.

    Prioridade:
    1. Campo turno vindo do sistema/banco.
    2. Horário de início.

    Isso resolve o caso mais comum do sistema: o Turno B fica salvo na DATA
    do dia produtivo, por exemplo 27/05 com hora 21:00, mas na planilha ele
    precisa aparecer como 26/05 21:00 até 27/05 06:00.
    """
    turno_txt = _normalizar(apontamento.get("turno") or apontamento.get("nome_turno") or "")
    if "B" == turno_txt or "TURNO B" in turno_txt or turno_txt.endswith(" B"):
        return "B"
    if "A" == turno_txt or "TURNO A" in turno_txt or turno_txt.endswith(" A"):
        return "A"

    hora_ini = _parse_hora(str(apontamento.get("hora_inicio") or "00:00"))
    if hora_ini >= time(18, 0) or hora_ini < time(6, 0):
        return "B"
    if time(6, 0) <= hora_ini < time(18, 0):
        return "A"
    return None


def _datetime_inicio_export(apontamento: dict, data_iso_export: str) -> datetime:
    """Data/hora que deve aparecer no Excel, respeitando o dia produtivo."""
    data_ref = _parse_data(data_iso_export)
    hora_ini = _parse_hora(str(apontamento.get("hora_inicio") or "00:00"))
    turno = str(apontamento.get("_turno_export") or _identificar_turno_export(apontamento) or "").upper()

    if turno == "B":
        # Turno B do dia 27 começa no dia 26 às 21:00.
        if hora_ini >= time(18, 0):
            return datetime.combine(data_ref - timedelta(days=1), hora_ini)
        return datetime.combine(data_ref, hora_ini)

    # Turno A sempre pertence ao próprio dia selecionado.
    return datetime.combine(data_ref, hora_ini)


def _datetime_fim_export(apontamento: dict, data_iso_export: str) -> datetime:
    """Data/hora final que deve aparecer no Excel, respeitando o dia produtivo."""
    data_ref = _parse_data(data_iso_export)
    hora_fim = _parse_hora(str(apontamento.get("hora_fim") or "00:00"))
    inicio = _datetime_inicio_export(apontamento, data_iso_export)
    turno = str(apontamento.get("_turno_export") or _identificar_turno_export(apontamento) or "").upper()

    if turno == "B":
        # Fim do Turno B do dia 27 é 27/05 06:00.
        if hora_fim <= time(8, 0):
            return datetime.combine(data_ref, hora_fim)
        fim = datetime.combine((data_ref - timedelta(days=1)), hora_fim)
        if fim < inicio:
            fim += timedelta(days=1)
        return fim

    fim = datetime.combine(data_ref, hora_fim)
    if fim < inicio:
        fim += timedelta(days=1)
    return fim


def _coletar_apontamentos_dia_produtivo(data_iso: str, turno: Optional[str] = None) -> List[dict]:
    """Busca somente os apontamentos salvos na data selecionada.

    Correção aplicada:
    - O sistema NÃO busca mais registros da data anterior automaticamente.
    - O Turno B continua aparecendo primeiro e com a lógica visual de 21:00 até 06:00.
    - A data/hora exibida no Excel continua sendo ajustada para o Turno B,
      mas a origem dos dados fica limitada ao dia escolhido no filtro.

    Exemplo:
    ao gerar 28/05/2026, o Excel usa apenas apontamentos salvos como 28/05/2026.
    Assim, não puxa Turno A ou Turno B que foram lançados no dia 27/05/2026.
    """
    registros: List[dict] = []
    vistos = set()

    def adicionar(apt: dict, turno_export: str):
        chave = (
            apt.get("id"), apt.get("data"), apt.get("data_iso"), apt.get("turno"),
            apt.get("maquina"), apt.get("produto"), apt.get("operador"),
            apt.get("hora_inicio"), apt.get("hora_fim"), apt.get("quantidade"), apt.get("perda"),
        )
        if chave in vistos:
            return
        vistos.add(chave)
        novo = dict(apt)
        novo["_turno_export"] = turno_export
        novo["_data_iso_export"] = data_iso
        registros.append(novo)

    # Busca somente a data escolhida no filtro/exportação.
    # Isso evita puxar apontamentos completos do dia anterior.
    try:
        lista_data = db.listar_apontamentos_para_excel(data_iso, None)
    except TypeError:
        lista_data = db.listar_apontamentos_para_excel(data_iso)
    except Exception:
        lista_data = []

    for apt in lista_data:
        t = _identificar_turno_export(apt)
        if t in ("B", "A"):
            adicionar(apt, t)

    if turno:
        t_filtro = str(turno).upper().replace("TURNO", "").strip()
        registros = [r for r in registros if str(r.get("_turno_export") or "").upper() == t_filtro]

    ordem = {"B": 0, "A": 1}
    return sorted(
        registros,
        key=lambda a: (
            ordem.get(a.get("_turno_export"), 9),
            str(a.get("maquina") or ""),
            _datetime_inicio_export(a, data_iso),
        ),
    )

def _cell(ws, row: int, col: int, value=None, fill=None, font=None, align="center", border=True, number_format=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if fill:
        c.fill = _fill(fill)
    if font:
        c.font = font
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        c.border = BORDA_PRETA
    if number_format:
        c.number_format = number_format
    return c


def _merge(ws, ref: str, value=None, fill=None, font=None, align="center", border=True, number_format=None):
    ws.merge_cells(ref)
    c = ws[ref.split(":")[0]]
    c.value = value
    if fill:
        c.fill = _fill(fill)
    if font:
        c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        c.number_format = number_format
    if border:
        for row in ws[ref]:
            for cell in row:
                cell.border = BORDA_PRETA
    return c


def _configurar_pagina(ws):
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = None

    larguras = {
        "A": 15, "B": 32, "C": 18, "D": 15, "E": 15, "F": 15, "G": 15, "H": 16,
        "I": 18, "J": 12, "K": 12, "L": 12, "M": 10, "N": 10, "O": 12, "P": 22,
        "Q": 15, "R": 15, "S": 12, "T": 18,
    }
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 18

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _cabecalho_principal(ws, data_iso: str, turno: Optional[str]):
    data_ref = _parse_data(data_iso)
    _cell(ws, 1, 1, f"DATA : {data_ref.strftime('%d/%m/%Y')}", fill=AZUL_CLARO, font=_font(True, PRETO, 12), align="left")

    titulo_turno = "TODOS" if not turno else str(turno).upper().replace("TURNO", "").strip()
    if titulo_turno == "TODOS":
        titulo = "PRODUTIVIDADE POR MAQUINAS (TURNO ´´B`` E ´´A``)"
    else:
        titulo = f"PRODUTIVIDADE POR MAQUINAS (TURNO ´´{titulo_turno}``)"
    _merge(ws, "B1:T1", titulo, fill=AZUL_CLARO, font=_font(True, PRETO, 16))

    headers = [
        "MAQUINA", "PRODUTO", "OPERADOR", "QTD PRODUZIDA", "HORA\nINÍCIO", "HORA\nFIM",
        "HORAS PARADAS", "HORAS\nTRABALHADAS", "CAP. PRODUÇÃO\n(Hrs trabalhadas)", "PRODUT. %",
        "TOTAL HORA", "HORA/DIA", "MÊS", "DATA", "OPERAÇÃO", "MOTIVO PARADA",
        "HORAS \"ÚTEIS\"", "HORAS NO DIA", "HORA/DIA2", "CAP. PRODUÇÃO\n(Turno completo)",
    ]
    for idx, h in enumerate(headers, start=1):
        _cell(ws, 2, idx, h, fill=AZUL_CLARO, font=_font(True, PRETO, 9))


def _linha_separadora(ws, linha: int):
    for col in range(1, 21):
        _cell(ws, linha, col, None, fill=PRETO)
    ws.row_dimensions[linha].height = 20


def _horas_turno_completo(turno: str) -> float:
    # Segue o padrão visual do modelo: B = 7,80 e A = 8,80.
    return 7.8 if str(turno).upper() == "B" else 8.8


def _preencher_lancamentos(ws, apontamentos: List[dict]) -> int:
    linhas_b = [a for a in apontamentos if a.get("_turno_export") == "B"]
    linhas_a = [a for a in apontamentos if a.get("_turno_export") == "A"]

    linha = 3
    for idx_grupo, (nome_turno, grupo) in enumerate([("B", linhas_b), ("A", linhas_a)]):
        # Não escreve texto no separador para manter igual ao modelo antigo: apenas faixa preta.
        _linha_separadora(ws, linha)
        linha += 1

        for apt in grupo:
            produto = str(apt.get("produto") or "").strip().upper()
            qtd = int(apt.get("quantidade") or 0)
            minutos_trab = tempo_para_minutos(apt.get("tempo_produzindo") or "00:00")
            horas_dec = minutos_trab / 60 if minutos_trab else 0
            prod_hora = _prod_hora(produto)
            cap_hrs_trabalhadas = round(prod_hora * horas_dec) if prod_hora else 0
            produtividade = qtd / cap_hrs_trabalhadas if cap_hrs_trabalhadas else None
            inicio = _datetime_inicio_export(apt, apt.get("_data_iso_export") or datetime.now().strftime("%Y-%m-%d"))
            fim = _datetime_fim_export(apt, apt.get("_data_iso_export") or datetime.now().strftime("%Y-%m-%d"))
            horas_turno = _horas_turno_completo(nome_turno)
            cap_turno_completo = round(prod_hora * horas_turno) if prod_hora else 0

            valores = [
                apt.get("maquina"), produto, apt.get("operador"), qtd, inicio, fim,
                _tempo_excel(apt.get("tempo_parado") or "00:00"), _minutos_para_excel(minutos_trab),
                cap_hrs_trabalhadas, produtividade,
                0, horas_dec, "#REF!", "#REF!",
                apt.get("processo") or apt.get("operacao") or "", apt.get("motivo_parada") or "",
                0, time(7, 48) if nome_turno == "B" else time(8, 48), horas_turno, cap_turno_completo,
            ]
            for col, valor in enumerate(valores, start=1):
                _cell(ws, linha, col, valor, fill=BRANCO, font=_font(True if col in [1, 2, 3] else False, PRETO, 10))

            ws.cell(linha, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(linha, 5).number_format = "d/m/yy h:mm"
            ws.cell(linha, 6).number_format = "d/m/yy h:mm"
            ws.cell(linha, 7).number_format = "h:mm"
            ws.cell(linha, 8).number_format = "hh:mm:ss"
            ws.cell(linha, 9).number_format = "#,##0"
            ws.cell(linha, 10).number_format = "0.0%"
            ws.cell(linha, 11).number_format = "hh:mm:ss"
            ws.cell(linha, 12).number_format = "0.00"
            ws.cell(linha, 17).number_format = "hh:mm:ss"
            ws.cell(linha, 18).number_format = "hh:mm:ss"
            ws.cell(linha, 19).number_format = "0.00"
            ws.cell(linha, 20).number_format = "#,##0"

            if produtividade is not None:
                if produtividade >= 0.80:
                    ws.cell(linha, 10).fill = _fill(VERDE_CLARO)
                    ws.cell(linha, 10).font = _font(True, VERDE_TEXTO, 10)
                else:
                    ws.cell(linha, 10).fill = _fill(VERMELHO_CLARO)
                    ws.cell(linha, 10).font = _font(True, VERMELHO_TEXTO, 10)

            linha += 1

    return linha + 1


def _agrupar_pre_acabados(apontamentos: List[dict]) -> List[dict]:
    agrupado: Dict[Tuple[str, str], dict] = {}
    for apt in apontamentos:
        if not _eh_pre_acabado(apt):
            continue
        maquina = str(apt.get("maquina") or "").strip().upper()
        produto = str(apt.get("produto") or "").strip().upper()
        qtd = int(apt.get("quantidade") or 0)
        if not produto or qtd <= 0:
            continue
        chave = (maquina, produto)
        if chave not in agrupado:
            agrupado[chave] = {"maquina": maquina, "produto": produto, "qtd": 0, "peso": 0.0}
        agrupado[chave]["qtd"] += qtd
        agrupado[chave]["peso"] += qtd * _peso_unitario(produto)
    return sorted(agrupado.values(), key=lambda x: (x["maquina"], x["produto"]))


def _preencher_pre_acabados(ws, linha_inicio: int, apontamentos: List[dict]) -> Tuple[int, float]:
    itens = _agrupar_pre_acabados(apontamentos)
    linha = linha_inicio

    _merge(ws, f"A{linha}:J{linha}", "PRÉ -ACABADOS ", fill=AZUL_CLARO, font=_font(True, PRETO, 14))
    linha += 1
    _cell(ws, linha, 1, "MÁQUINAS", fill=AZUL_CLARO, font=_font(True))
    _merge(ws, f"B{linha}:C{linha}", "PRODUTOS", fill=AZUL_CLARO, font=_font(True))
    _merge(ws, f"D{linha}:E{linha}", "PEÇAS", fill=AZUL_CLARO, font=_font(True))
    _merge(ws, f"F{linha}:J{linha}", "PESO KGS", fill=AZUL_CLARO, font=_font(True))
    linha += 1

    total_qtd = 0
    total_peso = 0.0
    if not itens:
        itens = [{"maquina": "", "produto": "", "qtd": 0, "peso": 0.0}]

    for item in itens:
        total_qtd += int(item["qtd"] or 0)
        total_peso += float(item["peso"] or 0)
        _cell(ws, linha, 1, item["maquina"], fill=BRANCO, font=_font(True))
        _merge(ws, f"B{linha}:C{linha}", item["produto"], fill=BRANCO, font=_font(True))
        _merge(ws, f"D{linha}:E{linha}", item["qtd"], fill=BRANCO, font=_font(True), number_format="#,##0")
        _merge(ws, f"F{linha}:J{linha}", item["peso"], fill=BRANCO, font=_font(True), number_format="#,##0.000")
        linha += 1

    linha += 1
    _cell(ws, linha, 1, None, fill=BRANCO)
    _merge(ws, f"B{linha}:C{linha}", None, fill=BRANCO)
    _merge(ws, f"D{linha}:E{linha}", total_qtd, fill=AMARELO, font=_font(True), number_format="#,##0")
    _merge(ws, f"F{linha}:J{linha}", total_peso, fill=AMARELO, font=_font(True), number_format="#,##0.000")

    return linha + 1, total_peso


def _preencher_meta(ws, linha_inicio: int, realizado_kg: float) -> int:
    linha = linha_inicio
    perc = realizado_kg / META_DIA_KG if META_DIA_KG else 0

    _merge(ws, f"A{linha}:D{linha}", "META DIA", fill=AZUL_CLARO, font=_font(True, PRETO, 12))
    _merge(ws, f"E{linha}:E{linha+3}", "X", fill=BRANCO, font=_font(True, PRETO, 34))
    _merge(ws, f"F{linha}:H{linha}", "REALIZADO", fill=AZUL_CLARO, font=_font(True, PRETO, 12))
    _merge(ws, f"I{linha}:J{linha}", "%", fill=AZUL_CLARO, font=_font(True, PRETO, 18))

    _merge(ws, f"A{linha+1}:D{linha+3}", META_DIA_KG, fill=BRANCO, font=_font(True, PRETO, 24), number_format="#,##0")
    _merge(ws, f"F{linha+1}:H{linha+3}", realizado_kg, fill=BRANCO, font=_font(True, PRETO, 24), number_format="#,##0.000")
    _merge(ws, f"I{linha+1}:J{linha+3}", perc, fill=BRANCO, font=_font(True, PRETO, 24), number_format="0.00%")

    for rng in [f"A{linha}:D{linha+3}", f"E{linha}:E{linha+3}", f"F{linha}:H{linha+3}", f"I{linha}:J{linha+3}"]:
        for row in ws[rng]:
            for cell in row:
                cell.border = BORDA_MEDIA
    for r in range(linha, linha + 4):
        ws.row_dimensions[r].height = 30

    return linha + 5


def _agrupar_perdas(apontamentos: List[dict]) -> List[dict]:
    agrupado: Dict[str, dict] = {}
    for apt in apontamentos:
        produto = str(apt.get("produto") or "").strip().upper()
        perda = int(apt.get("perda") or apt.get("qtd_perda") or 0)
        if not produto or perda <= 0:
            continue
        if produto not in agrupado:
            agrupado[produto] = {"produto": produto, "qtd": 0, "peso": 0.0}
        agrupado[produto]["qtd"] += perda
        agrupado[produto]["peso"] += perda * _peso_unitario(produto)
    return sorted(agrupado.values(), key=lambda x: x["produto"])


def _preencher_perdas(ws, linha_inicio: int, apontamentos: List[dict]) -> int:
    itens = _agrupar_perdas(apontamentos)
    linha = linha_inicio

    _merge(ws, f"A{linha}:B{linha}", "PERCAS", fill=LARANJA, font=_font(True, PRETO, 12))
    _cell(ws, linha, 3, "QTD", fill=LARANJA, font=_font(True, PRETO, 12))
    _cell(ws, linha, 4, "PESO", fill=LARANJA, font=_font(True, PRETO, 12))
    linha += 1

    total_qtd = 0
    total_peso = 0.0
    for item in itens:
        total_qtd += item["qtd"]
        total_peso += item["peso"]
        _merge(ws, f"A{linha}:B{linha}", item["produto"], fill=BRANCO, font=_font(True), align="left")
        _cell(ws, linha, 3, item["qtd"], fill=BRANCO, font=_font(True), number_format="#,##0")
        _cell(ws, linha, 4, item["peso"], fill=BRANCO, font=_font(True), number_format="#,##0.000")
        linha += 1

    # Mantém algumas linhas vazias igual ao modelo antigo, mas sem atrapalhar o Power BI.
    linhas_minimas = 10
    preenchidas = len(itens)
    for _ in range(max(0, linhas_minimas - preenchidas)):
        _merge(ws, f"A{linha}:B{linha}", "", fill=BRANCO)
        _cell(ws, linha, 3, 0, fill=BRANCO, number_format="#,##0")
        _cell(ws, linha, 4, 0, fill=BRANCO, number_format="#,##0.000")
        linha += 1

    _merge(ws, f"A{linha}:B{linha}", "", fill=MARROM)
    _cell(ws, linha, 3, total_qtd, fill=MARROM, font=_font(True, PRETO, 12), number_format="#,##0")
    _cell(ws, linha, 4, total_peso, fill=MARROM, font=_font(True, PRETO, 12), number_format="#,##0.000")
    return linha + 1


def gerar_excel(data_iso: str, turno: Optional[str] = None, saida: Optional[str] = None) -> str:
    """Gera o Excel no padrão completo da aba BD_LANÇAMENTOS."""
    apontamentos = _coletar_apontamentos_dia_produtivo(data_iso, turno)

    wb = Workbook()
    ws = wb.active
    ws.title = "BD_LANÇAMENTOS"

    _configurar_pagina(ws)
    _cabecalho_principal(ws, data_iso, turno)
    proxima_linha = _preencher_lancamentos(ws, apontamentos)
    proxima_linha, realizado_kg = _preencher_pre_acabados(ws, proxima_linha, apontamentos)
    proxima_linha = _preencher_meta(ws, proxima_linha, realizado_kg)
    _preencher_perdas(ws, proxima_linha, apontamentos)

    # Filtro exatamente nas colunas principais A:T.
    ws.auto_filter.ref = "A2:T2"

    if saida is None:
        nome_data = data_iso.replace("-", "")
        nome_turno = _safe_nome_arquivo(turno or "Todos")
        saida = str(Path.cwd() / f"BD_LANCAMENTOS_{nome_data}_{nome_turno}.xlsx")

    saida_path = Path(saida)
    if saida_path.suffix.lower() != ".xlsx":
        saida_path = saida_path.with_suffix(".xlsx")
    saida_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida_path)
    return str(saida_path)
